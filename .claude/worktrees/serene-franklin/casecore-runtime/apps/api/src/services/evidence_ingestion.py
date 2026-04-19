"""
Evidence Ingestion Service — Universal File Intake Pipeline

Handles ANY evidence file type through a unified pipeline:
  1. HASH — SHA-256 for chain-of-custody integrity + deduplication
  2. CLASSIFY — Determine file type and extraction strategy
  3. EXTRACT — Pull content (text from docs, transcript from audio/video,
               OCR from images, metadata from all)
  4. TIMELINE — Tag with dates, durations, temporal references
  5. MAP — Map extracted content to COA burden elements
  6. REGISTER — Create evidence record with full provenance chain

Supported file types:
  - Video: mp4, mov, avi, mkv, webm → Whisper transcription + frame extraction
  - Audio: mp3, wav, m4a, ogg, flac → Whisper transcription
  - Documents: pdf, docx, doc, txt, rtf → Text extraction
  - Images: jpg, png, tiff, bmp → OCR + metadata (EXIF dates, GPS)
  - Email: eml, msg → Header parsing + body + attachments
  - Spreadsheets: xlsx, csv → Structured data extraction
  - Messages: screenshots, exports → OCR + pattern matching
"""

import hashlib
import json
import mimetypes
import os
import subprocess
import tempfile
from datetime import datetime, timezone
from enum import Enum
from pathlib import Path
from typing import Optional

from pydantic import BaseModel, Field
from src.utils.ids import new_id


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------

class EvidenceType(str, Enum):
    VIDEO = "video"
    AUDIO = "audio"
    DOCUMENT = "document"
    IMAGE = "image"
    EMAIL = "email"
    SPREADSHEET = "spreadsheet"
    MESSAGE = "message"
    UNKNOWN = "unknown"


class ChainOfCustodyEntry(BaseModel):
    """Immutable record of every action taken on evidence."""
    entry_id: str
    timestamp: datetime
    action: str  # "ingested", "hashed", "extracted", "classified", "mapped", "reviewed"
    actor: str  # "system", "attorney", "client"
    details: Optional[dict] = None
    hash_at_action: Optional[str] = None  # SHA-256 at time of action


class EvidenceMetadata(BaseModel):
    """Extracted metadata from evidence file."""
    file_name: str
    file_size_bytes: int
    file_type: EvidenceType
    mime_type: Optional[str] = None
    sha256_hash: str

    # Temporal metadata
    created_date: Optional[str] = None  # from file metadata
    modified_date: Optional[str] = None
    content_dates: list[str] = Field(default_factory=list)  # dates found in content
    duration_seconds: Optional[float] = None  # for audio/video

    # Spatial metadata
    gps_location: Optional[str] = None  # from EXIF
    location_references: list[str] = Field(default_factory=list)  # places mentioned in content

    # Source metadata
    author: Optional[str] = None
    device: Optional[str] = None
    software: Optional[str] = None


class ExtractedContent(BaseModel):
    """Content extracted from an evidence file."""
    content_id: str
    evidence_id: str
    content_type: str  # "transcript", "text", "ocr", "structured_data", "metadata"
    full_text: str
    segments: list[dict] = Field(default_factory=list)  # timestamped segments for a/v
    tables: list[dict] = Field(default_factory=list)  # for spreadsheets/PDFs with tables
    entities_detected: list[dict] = Field(default_factory=list)  # people, places, dates, orgs
    extracted_at: datetime


class TimelineTag(BaseModel):
    """A point on the case timeline derived from evidence."""
    tag_id: str
    evidence_id: str
    timestamp_reference: str  # the date/time this refers to
    precision: str  # "exact", "day", "month", "approximate"
    description: str
    source_in_content: Optional[str] = None  # where in the content this was found
    confidence: float = Field(default=0.5, ge=0.0, le=1.0)


class IngestedEvidence(BaseModel):
    """Complete ingested evidence record."""
    evidence_id: str
    case_id: str
    intake_id: Optional[str] = None

    # File info
    metadata: EvidenceMetadata
    storage_path: str  # where the file is stored
    original_filename: str

    # Extracted content
    extracted_content: Optional[ExtractedContent] = None

    # Timeline
    timeline_tags: list[TimelineTag] = Field(default_factory=list)

    # COA mapping
    burden_element_mappings: list[dict] = Field(default_factory=list)
    # Each: {"element_id": str, "strength": str, "relevant_excerpt": str}

    # Chain of custody
    chain_of_custody: list[ChainOfCustodyEntry] = Field(default_factory=list)

    # Status
    status: str = "ingested"  # "ingested", "extracted", "mapped", "reviewed", "admitted"
    ingested_at: datetime
    mapped_at: Optional[datetime] = None


# ---------------------------------------------------------------------------
# File type classification
# ---------------------------------------------------------------------------

FILE_TYPE_MAP = {
    # Video
    ".mp4": EvidenceType.VIDEO, ".mov": EvidenceType.VIDEO,
    ".avi": EvidenceType.VIDEO, ".mkv": EvidenceType.VIDEO,
    ".webm": EvidenceType.VIDEO, ".wmv": EvidenceType.VIDEO,
    ".flv": EvidenceType.VIDEO, ".m4v": EvidenceType.VIDEO,
    # Audio
    ".mp3": EvidenceType.AUDIO, ".wav": EvidenceType.AUDIO,
    ".m4a": EvidenceType.AUDIO, ".ogg": EvidenceType.AUDIO,
    ".flac": EvidenceType.AUDIO, ".aac": EvidenceType.AUDIO,
    ".wma": EvidenceType.AUDIO,
    # Documents
    ".pdf": EvidenceType.DOCUMENT, ".docx": EvidenceType.DOCUMENT,
    ".doc": EvidenceType.DOCUMENT, ".txt": EvidenceType.DOCUMENT,
    ".rtf": EvidenceType.DOCUMENT, ".odt": EvidenceType.DOCUMENT,
    ".md": EvidenceType.DOCUMENT,
    # Images
    ".jpg": EvidenceType.IMAGE, ".jpeg": EvidenceType.IMAGE,
    ".png": EvidenceType.IMAGE, ".tiff": EvidenceType.IMAGE,
    ".tif": EvidenceType.IMAGE, ".bmp": EvidenceType.IMAGE,
    ".gif": EvidenceType.IMAGE, ".heic": EvidenceType.IMAGE,
    # Email
    ".eml": EvidenceType.EMAIL, ".msg": EvidenceType.EMAIL,
    # Spreadsheets
    ".xlsx": EvidenceType.SPREADSHEET, ".xls": EvidenceType.SPREADSHEET,
    ".csv": EvidenceType.SPREADSHEET, ".tsv": EvidenceType.SPREADSHEET,
    # Messages (screenshots typically come as images, but export files)
    ".json": EvidenceType.MESSAGE, ".html": EvidenceType.MESSAGE,
}


def classify_file(file_path: str) -> EvidenceType:
    ext = Path(file_path).suffix.lower()
    return FILE_TYPE_MAP.get(ext, EvidenceType.UNKNOWN)


# ---------------------------------------------------------------------------
# Evidence Ingestion Service
# ---------------------------------------------------------------------------

class EvidenceIngestionService:
    """
    Universal evidence ingestion pipeline.

    Handles any file type through:
      HASH → CLASSIFY → EXTRACT → TIMELINE → MAP → REGISTER
    """

    def __init__(
        self,
        storage_base: str = "/tmp/casecore_evidence",
        whisper_service=None,
        llm_provider=None,
        burden_tracker=None,
    ):
        self.storage_base = storage_base
        self.whisper = whisper_service
        self.llm = llm_provider
        self.tracker = burden_tracker

        # In-memory store (replace with persistence)
        self._evidence: dict[str, IngestedEvidence] = {}

        os.makedirs(storage_base, exist_ok=True)

    # ------------------------------------------------------------------
    # Main pipeline
    # ------------------------------------------------------------------

    def ingest_file(
        self,
        file_path: str,
        case_id: str,
        intake_id: Optional[str] = None,
        original_filename: Optional[str] = None,
        actor: str = "system",
    ) -> IngestedEvidence:
        """
        Full ingestion pipeline for a single file.

        Returns a complete IngestedEvidence record with hash,
        extracted content, timeline tags, and burden element mappings.
        """
        now = datetime.now(timezone.utc)
        evidence_id = new_id()
        fname = original_filename or Path(file_path).name

        # 1. HASH
        sha256 = self._compute_hash(file_path)

        # 2. CLASSIFY
        evidence_type = classify_file(file_path)
        mime = mimetypes.guess_type(file_path)[0] or "application/octet-stream"
        file_size = os.path.getsize(file_path)

        # 3. EXTRACT metadata
        file_metadata = self._extract_file_metadata(file_path, evidence_type)

        metadata = EvidenceMetadata(
            file_name=fname,
            file_size_bytes=file_size,
            file_type=evidence_type,
            mime_type=mime,
            sha256_hash=sha256,
            created_date=file_metadata.get("created_date"),
            modified_date=file_metadata.get("modified_date"),
            duration_seconds=file_metadata.get("duration"),
            author=file_metadata.get("author"),
            device=file_metadata.get("device"),
        )

        # 4. Store file
        storage_dir = os.path.join(self.storage_base, case_id)
        os.makedirs(storage_dir, exist_ok=True)
        storage_path = os.path.join(storage_dir, f"{evidence_id}_{fname}")

        import shutil
        shutil.copy2(file_path, storage_path)

        # Chain of custody: ingested
        chain = [
            ChainOfCustodyEntry(
                entry_id=new_id(),
                timestamp=now,
                action="ingested",
                actor=actor,
                details={"original_path": file_path, "storage_path": storage_path},
                hash_at_action=sha256,
            ),
            ChainOfCustodyEntry(
                entry_id=new_id(),
                timestamp=now,
                action="hashed",
                actor="system",
                details={"algorithm": "SHA-256", "hash": sha256},
                hash_at_action=sha256,
            ),
        ]

        # 5. EXTRACT content
        extracted = self._extract_content(
            file_path=storage_path,
            evidence_id=evidence_id,
            evidence_type=evidence_type,
            case_id=case_id,
        )

        if extracted:
            chain.append(ChainOfCustodyEntry(
                entry_id=new_id(),
                timestamp=datetime.now(timezone.utc),
                action="extracted",
                actor="system",
                details={"content_type": extracted.content_type, "text_length": len(extracted.full_text)},
                hash_at_action=sha256,
            ))

        # 6. TIMELINE tagging
        timeline_tags = []
        if extracted:
            timeline_tags = self._extract_timeline(
                extracted_content=extracted,
                evidence_id=evidence_id,
                file_metadata=file_metadata,
            )

        # 7. Build evidence record
        evidence = IngestedEvidence(
            evidence_id=evidence_id,
            case_id=case_id,
            intake_id=intake_id,
            metadata=metadata,
            storage_path=storage_path,
            original_filename=fname,
            extracted_content=extracted,
            timeline_tags=timeline_tags,
            chain_of_custody=chain,
            status="extracted" if extracted else "ingested",
            ingested_at=now,
        )

        # 8. MAP to burden elements (if tracker available)
        if self.tracker and extracted and extracted.full_text:
            mappings = self._map_to_burden_elements(case_id, evidence_id, extracted.full_text)
            evidence.burden_element_mappings = mappings
            if mappings:
                evidence.status = "mapped"
                evidence.mapped_at = datetime.now(timezone.utc)
                chain.append(ChainOfCustodyEntry(
                    entry_id=new_id(),
                    timestamp=datetime.now(timezone.utc),
                    action="mapped",
                    actor="system",
                    details={"elements_mapped": len(mappings)},
                    hash_at_action=sha256,
                ))

        # Store
        self._evidence[evidence_id] = evidence
        return evidence

    def ingest_bytes(
        self,
        data: bytes,
        filename: str,
        case_id: str,
        intake_id: Optional[str] = None,
        actor: str = "system",
    ) -> IngestedEvidence:
        """Ingest raw bytes (from upload) through the full pipeline."""
        with tempfile.NamedTemporaryFile(
            suffix=Path(filename).suffix, delete=False
        ) as tmp:
            tmp.write(data)
            tmp_path = tmp.name

        try:
            return self.ingest_file(tmp_path, case_id, intake_id, filename, actor)
        finally:
            os.unlink(tmp_path)

    def get_evidence(self, evidence_id: str) -> Optional[IngestedEvidence]:
        return self._evidence.get(evidence_id)

    def list_evidence(self, case_id: str) -> list[IngestedEvidence]:
        return [e for e in self._evidence.values() if e.case_id == case_id]

    # ------------------------------------------------------------------
    # Hash
    # ------------------------------------------------------------------

    def _compute_hash(self, file_path: str) -> str:
        """Compute SHA-256 hash for chain-of-custody integrity."""
        h = hashlib.sha256()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                h.update(chunk)
        return h.hexdigest()

    # ------------------------------------------------------------------
    # Metadata extraction
    # ------------------------------------------------------------------

    def _extract_file_metadata(self, file_path: str, evidence_type: EvidenceType) -> dict:
        """Extract file-level metadata (dates, duration, EXIF, etc.)."""
        meta = {}
        stat = os.stat(file_path)
        meta["modified_date"] = datetime.fromtimestamp(stat.st_mtime).isoformat()

        # For video/audio — try ffprobe for duration
        if evidence_type in (EvidenceType.VIDEO, EvidenceType.AUDIO):
            try:
                result = subprocess.run(
                    ["ffprobe", "-v", "quiet", "-print_format", "json",
                     "-show_format", "-show_streams", file_path],
                    capture_output=True, text=True, timeout=30,
                )
                if result.returncode == 0:
                    probe = json.loads(result.stdout)
                    fmt = probe.get("format", {})
                    meta["duration"] = float(fmt.get("duration", 0))
                    tags = fmt.get("tags", {})
                    meta["created_date"] = tags.get("creation_time")
                    meta["device"] = tags.get("encoder")
            except Exception:
                pass

        # For images — try to extract EXIF
        if evidence_type == EvidenceType.IMAGE:
            try:
                from PIL import Image
                from PIL.ExifTags import TAGS
                img = Image.open(file_path)
                exif = img._getexif()
                if exif:
                    for tag_id, value in exif.items():
                        tag = TAGS.get(tag_id, tag_id)
                        if tag == "DateTimeOriginal":
                            meta["created_date"] = str(value)
                        elif tag == "Make":
                            meta["device"] = str(value)
                        elif tag == "GPSInfo":
                            meta["gps"] = str(value)
            except Exception:
                pass

        return meta

    # ------------------------------------------------------------------
    # Content extraction
    # ------------------------------------------------------------------

    def _extract_content(
        self,
        file_path: str,
        evidence_id: str,
        evidence_type: EvidenceType,
        case_id: str,
    ) -> Optional[ExtractedContent]:
        """Extract text/content from evidence file based on type."""

        extractors = {
            EvidenceType.VIDEO: self._extract_av_content,
            EvidenceType.AUDIO: self._extract_av_content,
            EvidenceType.DOCUMENT: self._extract_document_content,
            EvidenceType.IMAGE: self._extract_image_content,
            EvidenceType.SPREADSHEET: self._extract_spreadsheet_content,
            EvidenceType.EMAIL: self._extract_email_content,
        }

        extractor = extractors.get(evidence_type)
        if not extractor:
            return None

        try:
            text, segments, extras = extractor(file_path, case_id)
            return ExtractedContent(
                content_id=new_id(),
                evidence_id=evidence_id,
                content_type=self._content_type_for(evidence_type),
                full_text=text,
                segments=segments,
                extracted_at=datetime.now(timezone.utc),
                **extras,
            )
        except Exception:
            return None

    def _extract_av_content(self, file_path: str, case_id: str) -> tuple[str, list, dict]:
        """Extract content from audio/video via Whisper transcription."""
        if not self.whisper:
            return "", [], {}

        transcript = self.whisper.transcribe_file(
            file_path=file_path,
            case_id=case_id,
            session_id=new_id(),
        )

        segments = [
            {
                "start": s.start_time,
                "end": s.end_time,
                "text": s.text,
                "confidence": s.confidence,
            }
            for s in transcript.segments
        ]

        return transcript.full_text, segments, {}

    def _extract_document_content(self, file_path: str, case_id: str) -> tuple[str, list, dict]:
        """Extract text from document files."""
        ext = Path(file_path).suffix.lower()

        if ext == ".txt" or ext == ".md":
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                return f.read(), [], {}

        if ext == ".pdf":
            try:
                from pypdf import PdfReader
                reader = PdfReader(file_path)
                pages = []
                for i, page in enumerate(reader.pages):
                    text = page.extract_text() or ""
                    pages.append({"page": i + 1, "text": text})
                full_text = "\n\n".join(p["text"] for p in pages)
                return full_text, pages, {}
            except Exception:
                return "", [], {}

        if ext == ".docx":
            try:
                from docx import Document
                doc = Document(file_path)
                paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
                return "\n\n".join(paragraphs), [], {}
            except Exception:
                return "", [], {}

        return "", [], {}

    def _extract_image_content(self, file_path: str, case_id: str) -> tuple[str, list, dict]:
        """Extract text from images via OCR."""
        try:
            import pytesseract
            from PIL import Image
            img = Image.open(file_path)
            text = pytesseract.image_to_string(img)
            return text, [], {}
        except Exception:
            return "", [], {}

    def _extract_spreadsheet_content(self, file_path: str, case_id: str) -> tuple[str, list, dict]:
        """Extract data from spreadsheets."""
        ext = Path(file_path).suffix.lower()
        try:
            if ext == ".csv" or ext == ".tsv":
                import csv
                delimiter = "\t" if ext == ".tsv" else ","
                with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                    reader = csv.reader(f, delimiter=delimiter)
                    rows = list(reader)
                text = "\n".join([", ".join(row) for row in rows[:1000]])
                tables = [{"rows": len(rows), "columns": len(rows[0]) if rows else 0}]
                return text, [], {"tables": tables}

            if ext in (".xlsx", ".xls"):
                import openpyxl
                wb = openpyxl.load_workbook(file_path, data_only=True)
                all_text = []
                tables = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))
                    sheet_text = "\n".join(
                        [", ".join(str(c) if c else "" for c in row) for row in rows[:500]]
                    )
                    all_text.append(f"=== {sheet_name} ===\n{sheet_text}")
                    tables.append({"sheet": sheet_name, "rows": len(rows)})
                return "\n\n".join(all_text), [], {"tables": tables}
        except Exception:
            pass
        return "", [], {}

    def _extract_email_content(self, file_path: str, case_id: str) -> tuple[str, list, dict]:
        """
        Extract content from email files (.eml, .msg).

        Parses email headers (From, To, Subject, Date) and body text.
        Attachments are noted but not recursively extracted (future enhancement).
        """
        ext = Path(file_path).suffix.lower()

        if ext == ".eml":
            try:
                import email
                from email import policy
                with open(file_path, "rb") as f:
                    msg = email.message_from_binary_file(f, policy=policy.default)

                # Extract headers
                headers = {
                    "from": str(msg.get("From", "")),
                    "to": str(msg.get("To", "")),
                    "cc": str(msg.get("Cc", "")),
                    "subject": str(msg.get("Subject", "")),
                    "date": str(msg.get("Date", "")),
                    "message_id": str(msg.get("Message-ID", "")),
                }

                # Extract body text
                body_parts = []
                if msg.is_multipart():
                    for part in msg.walk():
                        content_type = part.get_content_type()
                        if content_type == "text/plain":
                            payload = part.get_content()
                            if isinstance(payload, str):
                                body_parts.append(payload)
                        elif content_type == "text/html" and not body_parts:
                            # Fallback to HTML if no plain text
                            payload = part.get_content()
                            if isinstance(payload, str):
                                # Strip HTML tags for plain text extraction
                                import re
                                clean = re.sub(r'<[^>]+>', ' ', payload)
                                clean = re.sub(r'\s+', ' ', clean).strip()
                                body_parts.append(clean)
                else:
                    payload = msg.get_content()
                    if isinstance(payload, str):
                        body_parts.append(payload)

                body = "\n\n".join(body_parts)

                # Note attachments
                attachments = []
                if msg.is_multipart():
                    for part in msg.walk():
                        fn = part.get_filename()
                        if fn:
                            attachments.append({
                                "filename": fn,
                                "content_type": part.get_content_type(),
                                "size": len(part.get_payload(decode=True) or b""),
                            })

                # Build structured text
                full_text = (
                    f"EMAIL\n"
                    f"From: {headers['from']}\n"
                    f"To: {headers['to']}\n"
                    f"{'CC: ' + headers['cc'] + chr(10) if headers['cc'] else ''}"
                    f"Date: {headers['date']}\n"
                    f"Subject: {headers['subject']}\n"
                    f"{'Attachments: ' + ', '.join(a['filename'] for a in attachments) + chr(10) if attachments else ''}"
                    f"\n{body}"
                )

                segments = [{
                    "type": "email_header",
                    "from": headers["from"],
                    "to": headers["to"],
                    "subject": headers["subject"],
                    "date": headers["date"],
                }]

                extras = {
                    "tables": [{"type": "email", "attachments": len(attachments)}],
                }

                return full_text, segments, extras

            except Exception as e:
                logger.error(f"Email extraction failed for {file_path}: {e}")
                return "", [], {}

        if ext == ".msg":
            # .msg files require the extract-msg library
            try:
                import extract_msg
                msg = extract_msg.Message(file_path)
                headers = {
                    "from": msg.sender or "",
                    "to": msg.to or "",
                    "cc": msg.cc or "",
                    "subject": msg.subject or "",
                    "date": str(msg.date) if msg.date else "",
                }
                body = msg.body or ""
                attachments = [
                    {"filename": a.longFilename or a.shortFilename, "size": len(a.data or b"")}
                    for a in msg.attachments
                ] if msg.attachments else []

                full_text = (
                    f"EMAIL\n"
                    f"From: {headers['from']}\n"
                    f"To: {headers['to']}\n"
                    f"{'CC: ' + headers['cc'] + chr(10) if headers['cc'] else ''}"
                    f"Date: {headers['date']}\n"
                    f"Subject: {headers['subject']}\n"
                    f"{'Attachments: ' + ', '.join(a['filename'] for a in attachments) + chr(10) if attachments else ''}"
                    f"\n{body}"
                )
                msg.close()

                return full_text, [], {}

            except ImportError:
                logger.warning("extract-msg not installed. Run: pip install extract-msg")
                # Try to read as binary and extract what we can
                try:
                    with open(file_path, "rb") as f:
                        raw = f.read()
                    # Extract any readable ASCII strings
                    import re
                    strings = re.findall(rb'[\x20-\x7e]{20,}', raw)
                    text = "\n".join(s.decode("ascii", errors="replace") for s in strings[:100])
                    return f"[MSG file — partial extraction]\n{text}", [], {}
                except Exception:
                    return "", [], {}
            except Exception as e:
                logger.error(f"MSG extraction failed for {file_path}: {e}")
                return "", [], {}

        return "", [], {}

    def _content_type_for(self, evidence_type: EvidenceType) -> str:
        return {
            EvidenceType.VIDEO: "transcript",
            EvidenceType.AUDIO: "transcript",
            EvidenceType.DOCUMENT: "text",
            EvidenceType.IMAGE: "ocr",
            EvidenceType.SPREADSHEET: "structured_data",
        }.get(evidence_type, "unknown")

    # ------------------------------------------------------------------
    # Timeline extraction
    # ------------------------------------------------------------------

    def _extract_timeline(
        self,
        extracted_content: ExtractedContent,
        evidence_id: str,
        file_metadata: dict,
    ) -> list[TimelineTag]:
        """Extract timeline tags from content and metadata."""
        tags = []

        # Tag from file metadata dates
        if file_metadata.get("created_date"):
            tags.append(TimelineTag(
                tag_id=new_id(),
                evidence_id=evidence_id,
                timestamp_reference=file_metadata["created_date"],
                precision="exact",
                description="File creation date",
                confidence=0.9,
            ))

        # Use LLM to extract dates from content text
        if self.llm and extracted_content.full_text:
            try:
                system_prompt = """Extract all date and time references from this text.
Return valid JSON: {"dates": [{"reference": "date string", "precision": "exact|day|month|year|approximate", "description": "what happened", "confidence": 0.0-1.0}]}"""

                # Only send first 3000 chars to avoid token bloat
                raw = self.llm.complete(
                    system_prompt=system_prompt,
                    user_message=extracted_content.full_text[:3000],
                    max_tokens=1024,
                )
                parsed = json.loads(raw) if isinstance(raw, str) else raw
                for d in parsed.get("dates", []):
                    tags.append(TimelineTag(
                        tag_id=new_id(),
                        evidence_id=evidence_id,
                        timestamp_reference=d.get("reference", ""),
                        precision=d.get("precision", "approximate"),
                        description=d.get("description", ""),
                        confidence=d.get("confidence", 0.5),
                    ))
            except Exception:
                pass

        return tags

    # ------------------------------------------------------------------
    # Burden element mapping
    # ------------------------------------------------------------------

    def _map_to_burden_elements(
        self, case_id: str, evidence_id: str, text: str
    ) -> list[dict]:
        """Map extracted content to COA burden elements."""
        if not self.tracker or not self.llm:
            return []

        # Use burden tracker's map_response_to_elements
        updates = self.tracker.map_response_to_elements(case_id, text, self.llm)

        # Enrich with evidence_id
        mappings = []
        for u in updates:
            mappings.append({
                "element_id": u.get("element_id"),
                "evidence_id": evidence_id,
                "strength": u.get("strength"),
                "facts": u.get("facts", []),
                "notes": u.get("notes"),
            })
        return mappings
